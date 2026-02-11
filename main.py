"""
app.py - Processador de Relatórios SISU - IQ/UFF
Versão SEM pandas/SEM numpy - 100% compatível
"""

import streamlit as st
from openpyxl import load_workbook
from io import BytesIO
from datetime import datetime
import csv
import re

# Configuração da página - PRIMEIRO COMANDO
st.set_page_config(
    page_title="Processador SISU - IQ/UFF",
    page_icon="🧪",
    layout="wide"
)

# CSS
st.markdown("""
<style>
    .main-title {
        background: linear-gradient(90deg, #1e3c72, #2a5298);
        padding: 20px;
        border-radius: 10px;
        color: white;
        margin-bottom: 20px;
    }
    .metric-card {
        background-color: white;
        padding: 15px;
        border-radius: 10px;
        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        text-align: center;
    }
    .success-box {
        background-color: #d4edda;
        border-left: 5px solid #28a745;
        padding: 15px;
        border-radius: 5px;
    }
</style>
""", unsafe_allow_html=True)

def ler_arquivo_excel(uploaded_file):
    """Lê arquivo Excel sem pandas"""
    try:
        wb = load_workbook(uploaded_file, read_only=True, data_only=True)
        ws = wb.active
        
        dados = []
        headers = []
        
        # Pular linhas iniciais até encontrar cabeçalho
        start_row = 0
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if row and any('Situação' in str(cell) or 'Modalidade' in str(cell) for cell in row if cell):
                start_row = row_idx
                headers = [str(cell).strip() if cell else '' for cell in row]
                break
        
        if not headers:
            headers = [str(cell) if cell else f'Col{i}' for i, cell in enumerate(next(ws.iter_rows(values_only=True)), 1)]
            start_row = 1
        
        # Ler dados
        for row in ws.iter_rows(min_row=start_row + 1, values_only=True):
            if row and any(cell for cell in row):
                row_dict = {}
                for i, cell in enumerate(row):
                    if i < len(headers):
                        col_name = headers[i].lower()
                        row_dict[col_name] = str(cell).strip() if cell is not None else ''
                dados.append(row_dict)
        
        wb.close()
        return dados, None
    except Exception as e:
        return None, str(e)

def extrair_colunas(dados):
    """Extrai colunas importantes dos dados"""
    if not dados:
        return {}
    
    colunas_encontradas = {
        'situacao': None,
        'modalidade': None,
        'desvinculado': None,
        'curso': None
    }
    
    # Pegar primeira linha como referência
    primeiro = dados[0]
    
    for col in primeiro.keys():
        col_lower = col.lower()
        
        if 'situação' in col_lower or 'situacao' in col_lower:
            colunas_encontradas['situacao'] = col
        elif 'modalidade' in col_lower:
            colunas_encontradas['modalidade'] = col
        elif 'desvinculado' in col_lower or 'data' in col_lower and 'desvinc' in col_lower:
            colunas_encontradas['desvinculado'] = col
        elif 'curso' in col_lower or 'titulação' in col_lower or 'titulacao' in col_lower:
            colunas_encontradas['curso'] = col
    
    return colunas_encontradas

def processar_dados(dados, periodo):
    """Processa os dados sem pandas"""
    
    resultados = {
        'total': 0,
        'ativos': 0,
        'cancelados': 0,
        'cancelados_periodo': 0,
        'trancados': 0,
        'formados': 0,
        'cursos': {}
    }
    
    colunas = extrair_colunas(dados)
    
    if not colunas['situacao'] or not colunas['modalidade']:
        return None, "Colunas de Situação ou Modalidade não encontradas"
    
    for linha in dados:
        # Pular linhas de resumo
        if 'alunos de' in linha.get(colunas['situacao'], '').lower() or 'total' in linha.get(colunas['situacao'], '').lower():
            continue
        
        situacao = linha.get(colunas['situacao'], '').lower()
        modalidade = linha.get(colunas['modalidade'], '').upper()
        curso = linha.get(colunas.get('curso', ''), 'BACHAREL').lower()
        desvinculado = linha.get(colunas.get('desvinculado', ''), '')
        
        resultados['total'] += 1
        
        # Classificar curso
        if 'licenciatura' in curso:
            curso_nome = 'Licenciatura Química'
        elif 'bacharel' in curso:
            if 'industrial' in curso or 'q industrial' in curso:
                curso_nome = 'Bacharel Q Industrial'
            else:
                curso_nome = 'Bacharel Química'
        else:
            curso_nome = 'Bacharel Química'
        
        # Classificar modalidade
        if modalidade.startswith('A'):
            modalidade_nome = 'AMPLA CONCORRÊNCIA'
        elif modalidade.startswith('L'):
            modalidade_nome = 'AÇÕES AFIRMATIVAS'
        else:
            modalidade_nome = 'OUTROS'
        
        # Chave para dicionário
        key = f"{curso_nome}|{modalidade_nome}"
        
        if key not in resultados['cursos']:
            resultados['cursos'][key] = {
                'curso': curso_nome,
                'modalidade': modalidade_nome,
                'ingressantes': 0,
                'ativos': 0,
                'cancelados': 0,
                'trancados': 0,
                'formados': 0
            }
        
        resultados['cursos'][key]['ingressantes'] += 1
        
        # Classificar situação
        if any(x in situacao for x in ['pendente', 'inscrito', 'concluinte']):
            resultados['ativos'] += 1
            resultados['cursos'][key]['ativos'] += 1
        
        if 'trancado' in situacao:
            resultados['trancados'] += 1
            resultados['cursos'][key]['trancados'] += 1
        
        if 'formado' in situacao:
            resultados['formados'] += 1
            resultados['cursos'][key]['formados'] += 1
        
        if 'cancelamento' in situacao:
            resultados['cancelados'] += 1
            resultados['cursos'][key]['cancelados'] += 1
            
            # Verificar se é do período
            periodo_formatado = periodo.replace('.', '/')
            if periodo_formatado in str(desvinculado) or periodo in str(desvinculado):
                resultados['cancelados_periodo'] += 1
    
    return resultados, None

def main():
    # Header
    st.markdown("""
    <div class="main-title">
        <h1 style="color: white; margin:0;">🧪 Processador de Relatórios SISU</h1>
        <p style="color: white; margin:0; opacity:0.9;">Instituto de Química - UFF</p>
        <p style="color: white; margin:0; opacity:0.7; font-size:14px;">Versão Ultra Leve - SEM pandas</p>
    </div>
    """, unsafe_allow_html=True)
    
    # Sidebar
    with st.sidebar:
        st.markdown("### 📁 Upload do Relatório")
        
        uploaded_file = st.file_uploader(
            "Carregar arquivo Excel",
            type=['xlsx', 'xls'],
            help="Arquivo com a listagem de alunos"
        )
        
        periodo = st.text_input(
            "Período (ex: 2025.1)",
            value="2025.1"
        )
        
        processar = st.button(
            "🚀 Processar",
            type="primary",
            use_container_width=True,
            disabled=uploaded_file is None
        )
        
        st.markdown("---")
        st.markdown("""
        ### 📌 Instruções
        1. Upload do relatório (.xlsx)
        2. Confirmar período
        3. Clicar em Processar
        4. Copiar os valores
        """)
    
    # Área principal
    if uploaded_file and processar:
        with st.spinner("🔄 Processando arquivo..."):
            try:
                # Ler arquivo
                dados, erro = ler_arquivo_excel(uploaded_file)
                
                if erro:
                    st.error(f"❌ Erro ao ler arquivo: {erro}")
                    return
                
                if not dados:
                    st.error("❌ Nenhum dado encontrado no arquivo")
                    return
                
                # Processar dados
                resultados, erro = processar_dados(dados, periodo)
                
                if erro:
                    st.error(f"❌ Erro no processamento: {erro}")
                    return
                
                # MÉTRICAS
                st.success("✅ Processamento concluído com sucesso!")
                
                col1, col2, col3, col4, col5 = st.columns(5)
                
                with col1:
                    st.markdown('<div class="metric-card">', unsafe_allow_html=True)
                    st.metric("Total Alunos", resultados['total'])
                    st.markdown('</div>', unsafe_allow_html=True)
                
                with col2:
                    st.markdown('<div class="metric-card">', unsafe_allow_html=True)
                    st.metric("Matrículas Ativas", resultados['ativos'])
                    st.markdown('</div>', unsafe_allow_html=True)
                
                with col3:
                    st.markdown('<div class="metric-card">', unsafe_allow_html=True)
                    st.metric("Cancelamentos (Total)", resultados['cancelados'])
                    st.markdown('</div>', unsafe_allow_html=True)
                
                with col4:
                    st.markdown('<div class="metric-card">', unsafe_allow_html=True)
                    st.metric(f"Cancelamentos {periodo}", resultados['cancelados_periodo'])
                    st.markdown('</div>', unsafe_allow_html=True)
                
                with col5:
                    st.markdown('<div class="metric-card">', unsafe_allow_html=True)
                    st.metric("Formados", resultados['formados'])
                    st.markdown('</div>', unsafe_allow_html=True)
                
                # TABELA DE CURSOS
                st.markdown("### 📊 Dados por Curso e Modalidade")
                
                # Converter para lista para exibição
                cursos_lista = []
                for key, dados_curso in resultados['cursos'].items():
                    cursos_lista.append([
                        dados_curso['curso'],
                        dados_curso['modalidade'],
                        dados_curso['ingressantes'],
                        dados_curso['ativos'],
                        dados_curso['trancados'],
                        dados_curso['cancelados'],
                        dados_curso['formados']
                    ])
                
                # Ordenar por curso
                cursos_lista.sort(key=lambda x: (x[0], x[1]))
                
                # Criar tabela HTML
                tabela_html = """
                <table style="width:100%; border-collapse: collapse; background-color: white;">
                    <thead>
                        <tr style="background-color: #2a5298; color: white;">
                            <th style="padding: 10px; text-align: left;">Curso</th>
                            <th style="padding: 10px; text-align: left;">Modalidade</th>
                            <th style="padding: 10px; text-align: right;">Ingressantes</th>
                            <th style="padding: 10px; text-align: right;">Ativos</th>
                            <th style="padding: 10px; text-align: right;">Trancados</th>
                            <th style="padding: 10px; text-align: right;">Cancelados</th>
                            <th style="padding: 10px; text-align: right;">Formados</th>
                        </tr>
                    </thead>
                    <tbody>
                """
                
                for i, row in enumerate(cursos_lista):
                    bg_color = '#f9f9f9' if i % 2 == 0 else 'white'
                    tabela_html += f"""
                        <tr style="background-color: {bg_color};">
                            <td style="padding: 8px; border: 1px solid #ddd;">{row[0]}</td>
                            <td style="padding: 8px; border: 1px solid #ddd;">{row[1]}</td>
                            <td style="padding: 8px; border: 1px solid #ddd; text-align: right;">{row[2]}</td>
                            <td style="padding: 8px; border: 1px solid #ddd; text-align: right;">{row[3]}</td>
                            <td style="padding: 8px; border: 1px solid #ddd; text-align: right;">{row[4]}</td>
                            <td style="padding: 8px; border: 1px solid #ddd; text-align: right;">{row[5]}</td>
                            <td style="padding: 8px; border: 1px solid #ddd; text-align: right;">{row[6]}</td>
                        </tr>
                    """
                
                tabela_html += "</tbody></table>"
                st.markdown(tabela_html, unsafe_allow_html=True)
                
                # CONSOLIDADO PARA PLANILHA
                st.markdown("### 🎯 Dados para Planilha de Evasão")
                
                total_ingressantes = sum(row[2] for row in cursos_lista)
                total_cancelamentos = resultados['cancelados_periodo']
                total_ativos = sum(row[3] + row[4] for row in cursos_lista)
                total_formados = sum(row[6] for row in cursos_lista)
                taxa_evasao = (total_cancelamentos / total_ingressantes * 100) if total_ingressantes > 0 else 0
                
                # Tabela de consolidado
                consolidado_html = f"""
                <table style="width:100%; border-collapse: collapse; background-color: white; margin-top: 10px;">
                    <thead>
                        <tr style="background-color: #28a745; color: white;">
                            <th style="padding: 10px; text-align: left;">Período</th>
                            <th style="padding: 10px; text-align: right;">Ingressantes</th>
                            <th style="padding: 10px; text-align: right;">Cancelamentos</th>
                            <th style="padding: 10px; text-align: right;">Matrículas Ativas</th>
                            <th style="padding: 10px; text-align: right;">Formados</th>
                            <th style="padding: 10px; text-align: right;">% Evasão</th>
                        </tr>
                    </thead>
                    <tbody>
                        <tr style="background-color: white;">
                            <td style="padding: 8px; border: 1px solid #ddd;"><strong>{periodo}</strong></td>
                            <td style="padding: 8px; border: 1px solid #ddd; text-align: right;">{total_ingressantes}</td>
                            <td style="padding: 8px; border: 1px solid #ddd; text-align: right;">{total_cancelamentos}</td>
                            <td style="padding: 8px; border: 1px solid #ddd; text-align: right;">{total_ativos}</td>
                            <td style="padding: 8px; border: 1px solid #ddd; text-align: right;">{total_formados}</td>
                            <td style="padding: 8px; border: 1px solid #ddd; text-align: right;">{taxa_evasao:.1f}%</td>
                        </tr>
                    </tbody>
                </table>
                """
                st.markdown(consolidado_html, unsafe_allow_html=True)
                
                # INSTRUÇÕES
                st.markdown(f"""
                <div class="success-box" style="margin-top: 20px;">
                    <h4 style="margin-top:0; color: #155724;">📋 COPIAR PARA PLANILHA PRINCIPAL</h4>
                    <p><strong style="font-size: 16px;">Período: {periodo}</strong></p>
                    <table style="width:100%; border-collapse: collapse;">
                        <tr>
                            <td style="padding: 8px; background-color: #d4edda; border: 1px solid #c3e6cb;"><strong>Ingressantes:</strong></td>
                            <td style="padding: 8px; background-color: white; border: 1px solid #c3e6cb; font-size: 16px;"><strong>{total_ingressantes}</strong></td>
                            <td style="padding: 8px; background-color: #d4edda; border: 1px solid #c3e6cb;">→ Linha "Total de Ingressantes"</td>
                        </tr>
                        <tr>
                            <td style="padding: 8px; background-color: #d4edda; border: 1px solid #c3e6cb;"><strong>Cancelamentos:</strong></td>
                            <td style="padding: 8px; background-color: white; border: 1px solid #c3e6cb; font-size: 16px;"><strong>{total_cancelamentos}</strong></td>
                            <td style="padding: 8px; background-color: #d4edda; border: 1px solid #c3e6cb;">→ Linha "TOTAL CANCELAMENTOS"</td>
                        </tr>
                        <tr>
                            <td style="padding: 8px; background-color: #d4edda; border: 1px solid #c3e6cb;"><strong>Matrículas Ativas:</strong></td>
                            <td style="padding: 8px; background-color: white; border: 1px solid #c3e6cb; font-size: 16px;"><strong>{total_ativos}</strong></td>
                            <td style="padding: 8px; background-color: #d4edda; border: 1px solid #c3e6cb;">→ Linha "TOTAL MATRIC. ATIVAS"</td>
                        </tr>
                        <tr>
                            <td style="padding: 8px; background-color: #d4edda; border: 1px solid #c3e6cb;"><strong>Formados:</strong></td>
                            <td style="padding: 8px; background-color: white; border: 1px solid #c3e6cb; font-size: 16px;"><strong>{total_formados}</strong></td>
                            <td style="padding: 8px; background-color: #d4edda; border: 1px solid #c3e6cb;">→ Linha "Alunos Formados"</td>
                        </tr>
                        <tr>
                            <td style="padding: 8px; background-color: #d4edda; border: 1px solid #c3e6cb;"><strong>% Evasão:</strong></td>
                            <td style="padding: 8px; background-color: white; border: 1px solid #c3e6cb; font-size: 16px;"><strong>{taxa_evasao:.1f}%</strong></td>
                            <td style="padding: 8px; background-color: #d4edda; border: 1px solid #c3e6cb;">→ Linha "% Cancelamento"</td>
                        </tr>
                    </table>
                </div>
                """, unsafe_allow_html=True)
                
            except Exception as e:
                st.error(f"❌ Erro: {str(e)}")
                st.info("""
                **Dicas:**
                - Verifique se o arquivo é um Excel válido
                - Confirme se as colunas 'Situação' e 'Modalidade' existem
                - Tente salvar o arquivo como .xlsx antes de enviar
                """)

if __name__ == "__main__":
    main()
